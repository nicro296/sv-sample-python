import requests
import pandas as pd
import numpy as np
import json
import copy
import datetime
import openpyxl
from bs4 import BeautifulSoup as bs
from openpyxl.styles.alignment import Alignment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

import Archetypes as Arches
from Deck import Deck
from Matchup import Matchups
from Matchup import MatchupsXlsx


class JcgBracket:
    '''
    bracket : 
        an example is sample/jcgBracketSample.json
    {
        apiUrl : str
        bracket : dict
        groups : list[{
            id : int
                code : str
                group number
            name : str
            rounds : list[]
            sioName : str
        }]
        myUsername : str
        myTeamId : int
        myGroupCode : ""
        isa : bool
        settings : dict
    }
      
    matches : list[
        each match data json list 
        https://sv.j-cg.com/api/competition/group/XXXXXX (X=number)
        an example is sample/jcgBracketMatchSample.json
    {
        succes : bool
        rounds : list[{
            startAt : str
            name : str
            matches : list[{
                winner : int
                code : int
                id : int
                bo : int
                blockId : int
                groupCode : str
                teams : list[{
                    id : int
                    nameAbbr : str
                    name : str
                    nameKana : None
                    nicename : str
                    logo : str
                    win : bool
                    lose : bool
                    lost : bool
                    draw : bool
                    default : bool
                    wins : int
                    users : list[]
                }]
            }]
        }]
    }]
        
    parameters
    ----------
    url_id : str
        12 characters in URL
        https://sv.j-cg.com/competition/ *url_id* /home
    '''

    def __init__(self, url_id:str):
        url_groups  = 'https://sv.j-cg.com/competition/'+url_id+'/bracket#/1'
        source = requests.get(url_groups).text
        soup = bs(source, 'lxml')
        allscr = soup.find_all('script')
        tjson = allscr[7].text
        cleanedjson = tjson[tjson.find('{"apiUrl":"'):]
        self.bracket = json.loads(cleanedjson)
    def get_group_ids(self):
        group_ids = []
        for group in self.bracket['groups']:
            group_ids.append(group['id'])
        return group_ids
    def set_matches(self):
        self.matches = []
        group_ids = self.get_group_ids()
        for group_id in group_ids:
            bracketjson = 'https://sv.j-cg.com/api/competition/group/' + str(group_id)
            response = requests.get(bracketjson)
            self.matches.append(response.json())

class JcgEntry:
    '''
    players : list[{
        JcgEntryPlayer
    }]

    parameters
    ----------
    url_id : str
        12 characters in URL
        https://sv.j-cg.com/competition/ *url_id* /home
    '''
    def __init__(self, url_id:str):
        self.players = []
        url_entries = 'https://sv.j-cg.com/competition/'+url_id+'/entries'
        html = requests.get(url_entries)
        soup = bs(html.content, "html.parser")
        all_script = soup.find_all('script')
        dljson = all_script[7].string
        cleanedjson = dljson[dljson.find('list'):dljson.find('listFiltered')]
        finaljson = cleanedjson.replace('list:','').strip()[:-1]
        data = json.loads(finaljson)
        for player in data:
            self.players.append(JcgEntryPlayer(player))
    def get_player(self,username):
        for jcgEntryPlayer in self.players:
            if int(username) == jcgEntryPlayer.username:
                return jcgEntryPlayer
        raise Exception('dont find username')
    def write_json(self,file_uri):
        output = {'players':[]}
        for jcgEntryPlayer in self.players:
            output['players'].append(jcgEntryPlayer.get_dict())
        with open(file_uri ,'w', encoding='utf-8') as f:
            output_json = json.dumps(output)
            print(output_json, file=f)
    def set_archetype(self):
        for jcgEntryPlayer in self.players:
            jcgEntryPlayer.set_archetype()

class JcgEntryPlayer:
    '''PlayerData
    player_name : str
    player_id : int
        7 digit number
    username : int
        6 digit number
    nicename : str
        24 characters [0-9,a-z,A-z]
    result : int
        0 : not elected
        1 : elected
    decks : list[Deck]

    parameters
    -----------
    player : dict{
        id : int
        name :str
        username : str
        nicename : str
        result : int
        channel : None
        avatar : str
        comment : str
        show : bool
        sv_decks : list[{
            clan : int
            hash : str
        }]
    }
    '''
    def __init__(self,player):
        self.player_name = player['name']
        self.player_id = player['id']
        self.username = int(player['username'])
        self.nicename = player['nicename']
        self.result = player['result']
        self.decks = []
        for sv_deck in player['sv_decks']:
            deck = Deck(clan=sv_deck['clan'],hash=sv_deck['hash'])
            self.decks.append(deck)
    def set_archetype(self):
        for deck in self.decks:
            deck.set_archetype()
    def get_dict(self):
        output = copy.deepcopy(self)
        output.decks = []
        for deck in self.decks:
            output.decks.append(vars(deck))
        return vars(output)

'''サンプルjsonファイル作成用
# tournament_id = 'vsySSXEFrxP1'
# jcgBracket = JcgBracket(tournament_id)

# with open('sample/jcgBracketSample.json', 'w', encoding='utf-8') as f:
#     jb = json.dumps(jcgBracket.bracket)
#     print(jb, file=f)

# jcgBracket.set_matches()
# with open('sample/jcgBracketMatchSample.json' ,'w', encoding='utf-8') as f:
#     jb = json.dumps(jcgBracket.matches)
#     print(jb, file=f)

# jcgEntry = JcgEntry(tournament_id)
# jcgEntry.set_archetype()
# jcgEntry.write_json('sample/playersInJcgEntrySample.json')
'''


#分析パート 用意できていないので無理矢理くっつけておく

url_id = 'pvoF9tIfML42'
dt_now = datetime.datetime.now()
rule = ColorScaleRule(start_type='min', start_value=None, start_color='0000FF',
                        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                        end_type='max', end_value=None, end_color='FF0000')

Arch_Name = [
    [
        'テンポエルフ',
        'コントロールエルフ'
    ],
    [
        'コントロールロイヤル',
        'その他ロイヤル'
    ],
    [
        '秘術ウィッチ',
        'スペルウィッチ'
    ],
    [
        'コントロールドラゴン',
        'その他ドラゴン'
    ],
    [
        'ラスワ進化ネクロ',
        'ハイドラネクロ'
    ],
    [
        'ハンドレスヴァンパイア',
        '進化ヴァンパイア'
    ],
    [
        '結晶ビショップ',
        '回復ビショップ'
    ],
    [
        'AFネメシス',
        'その他（人形）'
    ]
    ]
def get_archetype_index(deck):
    for i in range(len(Arch_Name[(deck.clan-1)])):
        if Arch_Name[(deck.clan-1)][i] == deck.archetype_name:
            return i
    raise Exception('unvalid archetype name.')

DATA = [
    [[],[]], # CLASS_A,CLASS_B
    [[],[]], # WIN, LOSE
    [[],[]]  # arche_A,arche_B
    ]
'''
DATA[
    クラス情報:[
        対戦プレイヤーi(i=0or1):[j個目のデッキ]
    ],
    勝ち数情報:[
        対戦プレイヤーi(i=0or1):[j個目のデッキ]
    ],
    アーキタイプ番号情報:[
        対戦プレイヤーi(i=0or1):[j個目のデッキ]
    ]
]
'''
Head = ['デッキ1','デッキ２','勝率','データ数']
Head_SUM = ['デッキ','総合勝率','データ数']

class main():
    '''
    def make_DATA(self):
        for match in self.jcgBracket.matches:
            for _round in match['rounds']: 
                for round_match in _round['matches']:
                    for i in range(2): # i人目のデッキ
                        win = round_match['teams'][i]['wins']
                        player_id = round_match['teams'][i]['users'][0]['username']
                        player = self.jcgEntry.get_player(player_id)
                        if i == 0:
                            DATA[0][i].append(player.decks[0].clan)
                            DATA[0][i].append(player.decks[0].clan)
                            DATA[0][i].append(player.decks[1].clan)
                            DATA[0][i].append(player.decks[1].clan)
                            DATA[1][i].append(win)
                            DATA[1][i].append(win)
                            DATA[1][i].append(win)
                            DATA[1][i].append(win)
                            DATA[2][i].append(get_archetype_index(player.decks[0]))
                            DATA[2][i].append(get_archetype_index(player.decks[0]))
                            DATA[2][i].append(get_archetype_index(player.decks[1]))
                            DATA[2][i].append(get_archetype_index(player.decks[1]))
                        if i == 1:
                            DATA[0][i].append(player.decks[0].clan)
                            DATA[0][i].append(player.decks[1].clan)
                            DATA[0][i].append(player.decks[0].clan)
                            DATA[0][i].append(player.decks[1].clan)
                            DATA[1][i].append(win)
                            DATA[1][i].append(win)
                            DATA[1][i].append(win)
                            DATA[1][i].append(win)
                            DATA[2][i].append(get_archetype_index(player.decks[0]))
                            DATA[2][i].append(get_archetype_index(player.decks[1]))
                            DATA[2][i].append(get_archetype_index(player.decks[0]))
                            DATA[2][i].append(get_archetype_index(player.decks[1]))
    '''
    '''
    def anal(self):
        ANL_DATA = ([],[],[],[])
        #ANL_DATA = (
        #    デッキAのクラス:[],
        #    デッキBのクラス:[],
        #    勝率:[],
        #    デッキ数カウント:[]
        #)
        sum_win = 0 
        SUM_WIN = 0
        sum_lose = 0
        SUM_LOSE = 0
        SUM_Data = ([],[],[])
        #SUM_Data = (
        #    アーキタイプ名:[],
        #    勝率:[],
        #    デッキ数カウント:[]
        #)
        n = 0
        sum_n = 0
        wb = openpyxl.Workbook()
        for i in range(1,9): #デッキAのクラス
            for k in range(len(Arch_Name[i-1])): #デッキAのアーキタイプ 
                for j in range(1,9): #デッキBのクラス
                    for K in range(len(Arch_Name[j-1])): #デッキBのアーキタイプ
                        for N in range(len(DATA[0][0])): #N個目の
                            if DATA[0][0][N] == i and DATA[0][1][N] == j and DATA[2][0][N] == k and DATA[2][1][N] == K:#デッキとアーキタイプの照合
                                n = n + 1
                                sum_win = sum_win + int(DATA[1][0][N])
                                sum_lose = sum_lose + int(DATA[1][1][N])
                            if DATA[0][0][N] == j and DATA[0][1][N] == i and DATA[2][0][N] == K and DATA[2][1][N] == k:
                                n = n + 1
                                sum_win = sum_win + int(DATA[1][1][N])
                                sum_lose = sum_lose + int(DATA[1][0][N])
                        ANL_DATA[0].append(Arch_Name[i-1][k])
                        ANL_DATA[1].append(Arch_Name[j-1][K])
                        ANL_DATA[3].append(n)
                        if sum_win == 0:
                            ANL_DATA[2].append(0)
                        else:
                            ANL_DATA[2].append(round(100*sum_win/(sum_win+sum_lose),2)) #少数2桁
                        SUM_WIN = SUM_WIN + sum_win 
                        SUM_LOSE = SUM_LOSE + sum_lose
                        sum_n = sum_n + n
                        #変数の初期化 sum_win,sum_lose,n
                        sum_win = 0 
                        sum_lose = 0
                        n = 0
                SUM_Data[0].append(Arch_Name[i-1][k])
                if SUM_WIN == 0:
                    SUM_Data[1].append(0)
                else:
                    SUM_Data[1].append(round(100*SUM_WIN/(SUM_WIN+SUM_LOSE), 2))
                
                SUM_Data[2].append(round(sum_n))
                #変数の初期化 SUM_LOSE,SUM_WIN,sum_n
                SUM_LOSE = 0
                SUM_WIN = 0
                sum_n = 0
                #出力
                sheet = wb.create_sheet(ANL_DATA[0][0])
                SR = 2
                SC = 1
                # SR =
                for I in range(4):
                    sheet.cell(
                        row = 1,
                        column = I + 1 ,
                        value = Head[I]
                    ) 
                for I in range(len(ANL_DATA)):
                    for J in range(len(ANL_DATA[0])):
                        sheet.cell(
                            row = SR + J,
                            column = SC + I,
                            value = ANL_DATA[I][J] 
                            )
                for row in sheet:
                    for cell in row:
                        cell.alignment = Alignment(horizontal = 'center')
                sheet.column_dimensions['A'].width = 25
                sheet.column_dimensions['B'].width = 25
                sheet.conditional_formatting.add('C2:C17', rule)
                #初期化
                ANL_DATA = ([],[],[],[])
        ws = wb["Sheet"]
        for I in range(len(Head_SUM)):
            ws.cell(
                row = 1,
                column = I + 1 ,
                value = Head_SUM[I]
            ) 
        for i in range(len(SUM_Data[0])):
            ws.cell(
                row = 2 + i,
                column = 1,
                value = SUM_Data[0][i]
                )
            ws.cell(
                row = 2 + i,
                column = 2,
                value = SUM_Data[1][i]
                )
            ws.cell(
                row = 2 + i,
                column = 3,
                value = SUM_Data[2][i]
                )
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        for row in ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = 'center')
        ws.conditional_formatting.add('B2:B17', rule)
        wb = wb.save(str(dt_now.year)+str(dt_now.month).zfill(2)+str(dt_now.day).zfill(2)+'勝率.xlsx')
    '''
    def __main__(self):
        self.jcgBracket = JcgBracket(url_id)
        self.jcgBracket.set_matches()
        self.jcgEntry = JcgEntry(url_id)
        self.jcgEntry.set_archetype()
        # main.make_DATA()
        # main.anal()
        self.matchups = Matchups(Arches.archetypes)
        self.matchups.set_count(self.jcgBracket,self.jcgEntry)
        self.matchups_xlsx = MatchupsXlsx()
        archetype_name = 'テンポエルフ'
        self.matchups_xlsx.write_sheet(archetype_name, self.matchups.get_archetype_matchups(archetype_name))
        self.matchups_xlsx.save()

main = main()
if __name__ == '__main__':
    main.__main__()