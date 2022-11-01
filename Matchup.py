import openpyxl
import datetime
from openpyxl.styles.alignment import Alignment
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule


class Matchup:
    def __init__(self,archetype_main:str,archetype_opposite:str):
        self.archetype_main = archetype_main
        self.archetype_opposite = archetype_opposite
        self.win_main = 0
        self.win_opposite = 0
        self.count = 0
        if self.archetype_main == self.archetype_opposite:
            self.treshold = 2
            return
        self.treshold = 1
    def add_wins(self,archetype1,win1,archetype2,win2):
        self.count += self.treshold *1
        self.add_win(archetype1,win1)
        self.add_win(archetype2,win2)
    def add_win(self,archetype,win):
        if self.archetype_main == archetype:
            self.win_main += win
            return
        if self.archetype_opposite == archetype:
            self.win_opposite += win
            return
        raise Exception('unvalid archetype.')
    def exist(self,archetype):
        if archetype == self.archetype_main:
            return True
        if archetype == self.archetype_opposite:
            return True
        return False
    def get_opponent(self,archetype_main):
        if archetype_main == self.archetype_main:
            return self.archetype_opposite
        if archetype_main == self.archetype_opposite:
            return self.archetype_main
        print(archetype_main)
        raise Exception('unvalid archetype')
    def get_win_percentage(self,archetype):
        even = 50
        if archetype == self.archetype_main:
            if self.archetype_main == self.archetype_opposite:
                return even
            return round((self.win_main/(self.win_main+self.win_opposite))*100, 2)
        if archetype == archetype_opposite:
            return round((self.win_opposite/(self.win_main+self.win_opposite))*100, 2)
        raise Exception('unvali archetype')
    def get_data_count(self):
        return self.count

class Matchups:
    def __init__(self,archetypes):
        self.matchups = []
        for i in range(len(archetypes)):
            for j in range(len(archetypes)):
                if j < i:
                    continue
                self.matchups.append(Matchup(archetypes[i].archetype_name, archetypes[j].archetype_name))
    def set_count(self,jcg_bracket,jcg_entry):
        for group in jcg_bracket.matches:
            for _round in group['rounds']:
                for match in _round['matches']:
                    player1 = jcg_entry.get_player(match['teams'][0]['users'][0]['username'])
                    player2 = jcg_entry.get_player(match['teams'][1]['users'][0]['username'])
                    win1 = match['teams'][0]['wins']
                    win2 = match['teams'][1]['wins']
                    self.add_win(player1.decks[0].archetype_name, win1, player2.decks[0].archetype_name, win2)
                    self.add_win(player1.decks[0].archetype_name, win1, player2.decks[1].archetype_name, win2)
                    self.add_win(player1.decks[1].archetype_name, win1, player2.decks[0].archetype_name, win2)
                    self.add_win(player1.decks[1].archetype_name, win1, player2.decks[1].archetype_name, win2)
    def add_win(self,archetype1,win1,archetype2,win2):
        matchup = self.get_matchup(archetype1,archetype2)
        matchup.add_wins(archetype1,win1,archetype2,win2)
    def get_matchup(self,archetype1,archetype2):
        for matchup in self.matchups:
            if matchup.archetype_main == archetype1 and matchup.archetype_opposite == archetype2:
                return matchup
            if matchup.archetype_main == archetype2 and matchup.archetype_opposite == archetype1:
                return matchup
        print(archetype1+','+archetype2)
        raise Exception('unvlid archetype.')
    def get_archetype_matchups(self,archetype):
        output = []
        for match in self.matchups:
            if not match.exist(archetype):
                continue
            output.append(match)
        return output

dt_now = datetime.datetime.now()
Head = ['デッキ1','デッキ２','勝率','データ数']
Head_SUM = ['デッキ','総合勝率','データ数']
rule = ColorScaleRule(start_type='min', start_value=None, start_color='0000FF',
                        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                        end_type='max', end_value=None, end_color='FF0000')

class MatchupsXlsx:
    def __init__(self):
        self.wb = openpyxl.Workbook()
    def write_sheet(self,archetype,matchups):
        sheet = self.wb.create_sheet(archetype)
        for I in range(4):
            sheet.cell(
                row = 1,
                column = I + 1 ,
                value = Head[I]
            )
        row_count = 1
        for matchup in matchups:
            row_count += 1
            sheet.cell(
                row = row_count,
                column = 1,
                value = archetype
            )
            sheet.cell(
                row = row_count,
                column = 2,
                value = matchup.get_opponent(archetype)
            )
            sheet.cell(
                row = row_count,
                column = 3,
                value = matchup.get_win_percentage(archetype)
            )
            sheet.cell(
                row = row_count,
                column = 4,
                value = matchup.get_data_count()
            )
        for row in sheet:
            for cell in row:
                cell.alignment = Alignment(horizontal = 'center')
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 25
        sheet.conditional_formatting.add('C2:C17', rule)
    def ws(self):
        ws = self.wb["Sheet"]
        for I in range(len(Head_SUM)):
            ws.cell(
                row = 1,
                column = I + 1 ,
                value = Head_SUM[I]
            ) 
        for i in range(len(SUM_Data[0])):
            ws.cell(
                row = 2 + i,
                column = 1,
                value = SUM_Data[0][i]
                )
            ws.cell(
                row = 2 + i,
                column = 2,
                value = SUM_Data[1][i]
                )
            ws.cell(
                row = 2 + i,
                column = 3,
                value = SUM_Data[2][i]
                )
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        for row in ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = 'center')
        ws.conditional_formatting.add('B2:B17', rule)
    def save(self):
        self.wb.save(str(dt_now.year)+str(dt_now.month).zfill(2)+str(dt_now.day).zfill(2)+'勝率(sample).xlsx')